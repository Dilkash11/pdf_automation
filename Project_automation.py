
import PyPDF2
import re
import camelot
import pandas as pd
import os
import glob
import tabula
import openpyxl
 #GUI
from tkinter import *
from PIL import ImageTk, Image
#os.system("ghostscript-install.exe")

# Define global variables
pdf_folder_name = ""
excel_folder_name = ""
search_key = ""
#Function for GUI
def automate_extraction():
    # Get user inputs from GUI
    global pdf_folder_name, excel_folder_name, search_key
    pdf_folder_name = pdf_input.get()
    excel_folder_name = excel_input.get()
    search_key = keyword_search_entry.get()
    return(pdf_folder_name,excel_folder_name,search_key)

#GUI code starts here
root=Tk()
root.title("PDF_Automation")
root.geometry('500x500')
root.configure(background="black")
img=Image.open('th.jpg')
resized_img= img.resize((100,100))
img=ImageTk.PhotoImage(resized_img)

img_label=Label(root,image=img)
img_label.pack(pady=(10,10))

pdf_folder= Label(root, text="Enter the folder name where all the pdfs are stored", fg='white', bg='black')
pdf_folder.pack(pady=(20,5))
pdf_folder.config(font=('verdana',10))

pdf_var= StringVar()
pdf_input= Entry(root, width=60)
pdf_input.pack(ipady=4, pady=(1,15))

excel_folder= Label(root, text="Enter the folder name where you want all the output files to be stored", fg='white', bg='black')
excel_folder.pack(pady=(20,5))
excel_folder.config(font=('verdana',10))

excel_var= StringVar()
excel_input= Entry(root, width=60)
excel_input.pack(ipady=4, pady=(1,15))

keyword_search_label= Label(root, text="Please enter the search key", fg='white', bg='black')
keyword_search_label.pack(pady=(20,5))
keyword_search_label.config(font=('verdana',10))

keyword_search_var=StringVar()
keyword_search_entry= Entry(root, width=60)
keyword_search_entry.pack(ipady=4, pady=(1,15))

automate_button = Button(root, text="Automate", command=automate_extraction)
automate_button.pack(pady=(20, 10))

# Add a button to start the automation
start_button = Button(root, text="Start Automation", command=start_automation)
start_button.pack(pady=(20, 10))

root.mainloop()

#os.system("ghostscript-install.exe")


searchKey = search_key
sowFolderName = pdf_folder_name
excelFolderName = excel_folder_name
masterFileName = "Master_File.xlsx"
errorFileName = "File_error.xlsx"
sowListFileName = "Filename_List.xlsx"
columnHeaderSowListFile = "Filename"

currentDir= os.getcwd()
sowFolderPath= os.path.join(currentDir,sowFolderName)
excelFolderPath= os.path.join(currentDir, excelFolderName)

if not os.path.exists(excelFolderPath):
    os.makedirs(excelFolderPath)
    print(f"Folder '{excelFolderName}' created successfully in the current directory.")
else:
    print(f"Folder '{excelFolderName}' already exists in the current directory.")
allSowFiles = glob.glob(f"{sowFolderPath}/*.Pdf")
requiredSowFiles = []
errorFiles = []
masterFilePath = excelFolderPath + "//" + masterFileName
errorFilePath = excelFolderPath + "//" + errorFileName
sowListFilePath = excelFolderPath + "//" + sowListFileName

# Search keyword in the pdf file.def searchWord(pdfFile , keyword):
def searchWord(pdfFile , keyword):
    pdfReader = PyPDF2.PdfFileReader(pdfFile)
    totalPage = pdfReader.getNumPages()
    for i in range(totalPage):
        pageObject = pdfReader.getPage(i)
        pageData = pageObject.extractText()
        if re.findall(keyword, pageData):
          return str(i+1)

# Find the list of new SOW files for which data has to be extracted.
requiredSowFiles=allSowFiles
if(os.path.exists(sowListFilePath)):
    df = pd.read_excel(sowListFilePath)
    fileList = df[columnHeaderSowListFile].tolist()
    requiredSowFiles=[]
    for file in allSowFiles:
        if (os.path.basename(file).split(".")[0]) not in fileList:
          requiredSowFiles.append(file)
       
             
# Create individual excel files for each SOW Files.
for sow in requiredSowFiles:
  try:
    indexPage = searchWord(sow, searchKey)
    tables = camelot.read_pdf(sow, pages = indexPage)
    requiredTable= tables[0].df      
    # for i in range(int(indexPage)+1, totalPage):
    #   if ((tables[0].df[0][len(tables[0].df)-1]) == "Estimated Fees"):
    #     break      
    #   else:
    #     tables= camelot.read_pdf(sow, pages = str(i), flavor = "stream")
    #     requiredTable.append(tables[0].df) 

    
    sowFileName=os.path.basename(sow).split(".")[0]  
    requiredTable.to_excel(sowFolderPath+f"//{sowFileName}.xlsx", index = False)
  except Exception as e:
    errorFiles.append(sow)

# Create an excel file containing file name for which data extraction from SOW faulted.
# Caution: This deletes the previously generated error file
if len(errorFiles) != 0:
  df=pd.DataFrame(errorFiles)
  if(os.path.exists(errorFilePath)):
    os.remove(errorFilePath)
  df.to_excel(errorFilePath, index = False, header=False)  

# Now all the extracted tables from respective pdfs are individual excel file
 # All these files will be merged into one master file
sowExcelFiles = glob.glob(sowFolderPath+"/*.xlsx")
sowTableList = pd.DataFrame()
sowNameList=[]
newMasterFilePath = excelFolderPath + "//New_master_file.xlsx"  
# Store names of all SOW files for which data extraction was successful.


for file in sowExcelFiles:
  df=pd.read_excel(file)
  df["Filename"]= os.path.basename(file).split(".")[0]
  sowTableList= sowTableList.append(df, ignore_index=True)
  sowNameList.append(os.path.basename(file).split(".")[0]) 
# Concatenate all the tables present in the SOW table list and convert the dataframe into master excel file.
sowTableList.to_excel(newMasterFilePath, index=False)
# mergedTable = pd.concat(sowTableList,ignore_index=True)
# mergedTable.to_excel(newMasterFilePath, index = False)
# Create excel file with names of all SOW files for which data extraction was successful.

#Merge the new master file data with existing master file.
# Caution: Only one master file will remain with merged data after this operation.
if(os.path.exists(masterFilePath)):
  df1 = pd.read_excel(masterFilePath)
  df2 = pd.read_excel(newMasterFilePath)
  df = pd.concat([df1,df2],ignore_index=True)
  os.remove(masterFilePath)
  df.to_excel(masterFilePath,index=False)
  os.remove(newMasterFilePath)
else:
      os.rename(newMasterFilePath, masterFilePath)

#Merge the new SOW name list with existing list.
# Caution: Only one file will remain with merged data after this operation.
sowNameList = pd.DataFrame(sowNameList,columns=["Filename"])
if(os.path.exists(sowListFilePath)):
  df1 = pd.read_excel(sowListFilePath)
  df = pd.concat([df1, sowNameList],ignore_index=True)
  os.remove(sowListFilePath)
  df.to_excel(sowListFilePath,index=False)
else:
    sowNameList.to_excel(sowListFilePath,index=False,header=True,startrow=0)

#Make the value of cell"A1" in Sow_List.xlsx equals to "SowFilename"
df=pd.read_excel(sowListFilePath,header=0)
df.rename(columns={'Unnamed: 0': 'Filename'}, inplace=True)
df.to_excel(sowListFilePath, index=False, header=True, startrow=0)




# Delete all individual excel files created for each SOW.
for file in sowExcelFiles:
  os.remove(file)

# Automation of Master_file, to remove unnecessary information

from openpyxl import workbook, load_workbook
if(os.path.exists(masterFilePath)):
  wb=load_workbook(masterFilePath)
  ws=wb.active
  last_row=ws.max_row
  for i in range(3,last_row):
   if ws.cell(row=i,column= 1).value=='Role':
      ws.delete_rows(i)
  for i in range(2,last_row):
   if ws.cell(row=i,column= 1).value=='Estimated Fees':
      ws.delete_rows(i)

wb.save(masterFilePath)
# deleting duplicate rows if it is present
if(os.path.exists(masterFilePath)):
  df=pd.read_excel(masterFilePath)
   #Drop duplicate rows
  df.drop_duplicates(inplace=True)

# Save the dataframe back to the Excel file
df.to_excel(masterFilePath, index=False)


